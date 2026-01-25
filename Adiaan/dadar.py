import os
import requests
from datetime import datetime
from collections import Counter
from ethiopian_date import EthiopianDateConverter 

# Library-wwan mirkaneeffachuu
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from fpdf import FPDF
except ImportError:
    print("\n[!] Maaloo: 'pip install openpyxl requests fpdf ethiopian-date' godhaa.")

# --- QINDAA'INA (CONFIG) HAARAA ---
USER_NAME = "admin"
PASS_WORD = "1234"
BOT_TOKEN = "8357193631:AAHCuSnXzjZTQaglkmcS0gq-EvqnkIQLDBI"
CHAT_ID_MANAGER = "7329587700" 

# Ragaalee SMS Gateway
SMS_TOKEN = "7b96636f-e286-4aae-ba20-b7dd310897db"

# IP Address amma App Gateway kee irratti argitu kanaan bakka buusi
SMS_URL = "http://10.181.252.6:8082/send" 

DATA_FILE = "dadar_final_report.txt"
OFFICE_HEAD = "Obbo Aqiil Abdujaalil" 
LOGO_PATH = "logo.png" 

# --- FUNKSHIINIIWWAN GARGAARTUU ---

def guyyaa_itophiyaa(year, month, day):
    """Guyyaa G.C gara E.C jijjiira"""
    try:
        eth_year, eth_month, eth_day = EthiopianDateConverter.to_ethiopian(year, month, day)
        return f"{eth_day}/{eth_month}/{eth_year} E.C"
    except:
        return f"{day}/{month}/{year} G.C"

def send_sms(phone, message):
    """SMS Gateway kallaattiin akka erguuf (Device ID dabalatee)"""
    try:
        # Lakk bilbilaa sirreessuuf (+251...)
        if phone.startswith('0'):
            phone = "+251" + phone[1:]
        elif not phone.startswith('+'):
            phone = "+251" + phone
            
        payload = {
            'token': SMS_TOKEN,
            'device': DEVICE_ID, # Device ID amma asitti dabalameera
            'to': phone,
            'message': message
        }
        
        # Ergaa POST fayyadamnee ergina
        res = requests.post(SMS_URL, data=payload, timeout=10)
        
        if res.status_code == 200:
            print(f"[✓] SMS gara {phone} tti ergameera.")
            return True
        else:
            print(f"[!] Gateway Error: {res.status_code} - {res.text}")
            return False
    except Exception as e:
        print(f"[!] Network Error: {e}")
        return False

def send_telegram_text(message):
    """Ergaa barreeffamaa qofa qondaalaaf erga"""
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    try:
        requests.post(url, data={'chat_id': CHAT_ID_MANAGER, 'text': message, 'parse_mode': 'Markdown'}, timeout=20)
    except: pass

def send_telegram_file(file_path, caption=""):
    """Excel ykn PDF Telegram irratti erga"""
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"
    try:
        with open(file_path, 'rb') as f:
            requests.post(url, data={'chat_id': CHAT_ID_MANAGER, 'caption': caption}, files={'document': f}, timeout=20)
    except: pass

# --- KUTAA DOOKUMANTIIWWANII ---

def uumi_sartifiketii(ogeessa, rank, waggaa):
    """Sartifiketii PDF ogeessaaf qopheessa"""
    try:
        pdf = FPDF(orientation='L', unit='mm', format='A4')
        pdf.add_page()
        pdf.set_draw_color(31, 78, 120); pdf.set_line_width(1.5); pdf.rect(10, 10, 277, 190)
        
        if os.path.exists(LOGO_PATH):
            pdf.image(LOGO_PATH, x=128, y=15, w=40)
            pdf.ln(45)
        else: pdf.ln(25)

        pdf.set_font('Arial', 'B', 24)
        pdf.cell(0, 10, 'BULCHIINSA MAGAALAA DADAR', ln=True, align='C')
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 8, 'Wajjiira Lafa Bulchiinsa Magaalaa', ln=True, align='C')
        pdf.ln(10)
        
        pdf.set_font('Arial', 'B', 30); pdf.set_text_color(31, 78, 120)
        pdf.cell(0, 20, 'SARTIFIKETII BADHAASA WAGGAA', ln=True, align='C')
        
        pdf.set_text_color(0, 0, 0); pdf.set_font('Arial', '', 16)
        pdf.cell(0, 15, f"Sartifiketiin kun Ogeessa kabajamaa {ogeessa.upper()}f", ln=True, align='C')

txt = (f"tajaajila mamiilaa haala bareedaa fi quubsaa ta'een waggaa {waggaa} "
               f"kennaa turaniif badhaasa {rank} ta'uun qophaa'eef.")
        pdf.set_font('Arial', '', 14); pdf.multi_cell(0, 8, txt, align='C')
        pdf.ln(20)
        
        pdf.set_font('Arial', 'B', 14); pdf.cell(0, 8, f"{OFFICE_HEAD}", ln=True, align='C')
        pdf.cell(0, 8, "Itti Gaafatamaa Wajjiraa", ln=True, align='C')
        
        f_name = f"Sartifiketii_{ogeessa.replace(' ', '_')}.pdf"
        pdf.output(f_name)
        return f_name
    except: return None

def uumi_gabaasa_target(filter_type, value, label):
    """Excel uumee Telegram-itti erga"""
    if not os.path.exists(DATA_FILE): return
    
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Gabaasa"
    headers = ["Guyyaa", "Abbaa Dhimmaa", "Araddaa", "Wirtuu", "Gosa Tajaajilaa", "Ogeessa", "Beellama", "Waligala"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", start_color="1F4E78")
    
    total_rev, count, ogeessota = 0, 0, []
    with open(DATA_FILE, "r") as f:
        for line in f:
            p = line.strip().split("|")
            if len(p) < 11: continue
            dt = datetime.strptime(p[0], "%Y-%m-%d %H:%M")
            match = (filter_type == "guyyaa" and dt.strftime("%A") == value) or \
                    (filter_type == "jia" and dt.month == value) or \
                    (filter_type == "waggaa" and dt.year == value)
            if match:
                eth_d = guyyaa_itophiyaa(dt.year, dt.month, dt.day)
                ws.append([eth_d, p[1], p[2], p[3], p[5], p[6], p[8], p[10]])
                total_rev += float(p[10]); ogeessota.append(p[6]); count += 1

    if count > 0:
        f_name = f"Gabaasa_{label}.xlsx"
        wb.save(f_name)
        if filter_type == "waggaa":
            top_3 = Counter(ogeessota).most_common(3)
            ranks = ["1ffaa", "2ffaa", "3ffaa"]
            for i, (name, _) in enumerate(top_3):
                sf = uumi_sartifiketii(name, ranks[i], value)
                if sf: send_telegram_file(sf, f"📜 Badhaasa: {name}")
        
        caption = f"📊 {label}\n💰 Galii: {total_rev} ETB\n📑 Galmee: {count}"
        send_telegram_file(f_name, caption)
        print(f"[✓] Gabaasaan ergameera.")

# --- KUTAA HOJII ---

def galmeessi():
    print("\n" + "="*15 + " GALMEE HAARAA " + "="*15)
    ad = input("Maqaa Abbaa Dhimmaa: ")
    ar = input("Araddaa: ")
    wi = input("Wirtuu: ")
    bad = input("Bilbila AD: ")
    gs = input("Gosa Tajaajilaa: ")
    og = input("Maqaa Ogeessaa: ")
    bog = input("Bilbila Ogeessaa: ")
    gb = input("Guyyaa Beellamaa (YYYY-MM-DD): ") or "2026-01-01"
    sb = input("Sa'aatii Beellamaa: ") or "08:00"
    
    k_vals = []
    print("\n--- Kafaltiiwwan ---")
    for kt in ["Kartaa", "Itti Fayyadaama", "Jij_Maqaa", "Lizii_Dura", "TOT", "Gibira", "Kan_Biro"]:
        v = input(f"{kt}: ") or "0"
        try: k_vals.append(float(v))
        except: k_vals.append(0.0)
            
    waligala = sum(k_vals)
    now_s = datetime.now().strftime("%Y-%m-%d %H:%M")
    k_str = "|".join(map(str, k_vals))
    line = f"{now_s}|{ad}|{ar}|{wi}|{bad}|{gs}|{og}|{bog}|{gb} {sb}|{k_str}|{waligala}\n"
    
    with open(DATA_FILE, "a") as f: f.write(line)

    # 1. SMS Abbaa Dhimmaaf
    msg_ad = f"Kabajamaa {ad}, Araddaa {ar}-Wirtuu {wi} irratti tajaajila {gs}af galmeeffamtaniittu. Ogeessi: {og}. Beellama: {gb} {sb}. W/B/L/M/Dadar."
    send_sms(bad, msg_ad)

    # 2. SMS Ogeessaaf
    msg_og = f"Ogeessa {og}, tajaajilli {gs} kan mamiila {ad} (Araddaa: {ar}) isiniif kennameera. Beellama: {gb} {sb}."
    send_sms(bog, msg_og)

    # 3. Telegram Notification
    tel_msg = f"✅ *GALMEE HAARAA*\n👤 AD: {ad}\n📍 Bakka: {ar} | {wi}\n🛠 Tajaajila: {gs}\n💰 Waligala: {waligala} ETB\n🧑‍💼 Ogeessa: {og}\n📅 Beellama: {gb} {sb}"
    send_telegram_text(tel_msg)
    print("\n[✓] Galmeeffameera! SMS fi Notification ergameera.")

# --- NAVIGATION ---
